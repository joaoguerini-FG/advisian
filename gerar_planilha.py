"""Gera planilha Excel simples: processo | regra sugerida | integra da publicacao"""
import json, sys, re
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open("intimacoes_state.json", "r", encoding="utf-8") as f:
    state = json.load(f)

pubs = state["publicacoes"]
print(f"Total publicacoes: {len(pubs)}")

def limpar(texto):
    if not texto:
        return ""
    # Remover CSS
    t = re.sub(r"[#.]?[a-zA-Z][\w-]*\s*\{[^}]*\}", " ", texto)
    # Remover props CSS soltas
    t = re.sub(r"(font-family|padding|margin|color|background|border|display|width|height|line-height|text-align|font-size|font-weight|font-style)\s*:\s*[^;\n]+;?", " ", t, flags=re.I)
    # Remover HTML
    t = re.sub(r"<[^>]+>", " ", t)
    # Remover entities
    t = re.sub(r"&[a-z]+;", " ", t)
    t = re.sub(r"&#\d+;", " ", t)
    # Normalizar
    t = re.sub(r"\s+", " ", t).strip()
    return t

wb = Workbook()
ws = wb.active
ws.title = "Intimacoes DJEN"

# Cabecalho
headers = ["Processo", "Data", "Natureza", "Regra V1 (base historica)", "Regra V2 (ground truth)", "Confianca", "Prazo", "Mudou?", "Integra da Publicacao"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def color_regra(cell, regra):
    if "MANUAL" in (regra or "") or "NENHUMA" in (regra or ""):
        cell.fill = PatternFill("solid", fgColor="FED7AA")
        cell.font = Font(color="7C2D12", bold=True)
    elif regra == "INFORMATIVO_SEM_PRAZO":
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
        cell.font = Font(color="1E3A8A")
    elif regra and regra not in ("PENDENTE", "PENDENTE_CLASSIFICACAO", "ERRO_CLASSIFICACAO", ""):
        cell.font = Font(bold=True, color="4C1D95")

# Dados
for idx, pub in enumerate(pubs, start=2):
    cls = pub.get("classificacao", {})
    regra_v2 = cls.get("regra", "")
    regra_v1 = cls.get("regra_v1", regra_v2)  # Se nao tem v1, usa v2 (nao foi reclassificada)
    confianca = cls.get("confianca", "")
    prazo = cls.get("prazo_dias", "")
    processo = pub.get("processo", "")
    data = pub.get("data_disponibilizacao", "")
    natureza = pub.get("natureza", "") or pub.get("contexto", {}).get("natureza", "")
    texto = limpar(pub.get("texto_completo", "") or pub.get("texto_resumo", ""))

    ws.cell(row=idx, column=1, value=processo)
    ws.cell(row=idx, column=2, value=data)
    ws.cell(row=idx, column=3, value=natureza)

    # V1
    v1_cell = ws.cell(row=idx, column=4, value=regra_v1)
    color_regra(v1_cell, regra_v1)

    # V2
    v2_cell = ws.cell(row=idx, column=5, value=regra_v2)
    color_regra(v2_cell, regra_v2)

    # Confianca
    c_cell = ws.cell(row=idx, column=6, value=confianca)
    if confianca == "ALTA":
        c_cell.fill = PatternFill("solid", fgColor="C6F6D5")
        c_cell.font = Font(color="22543D", bold=True)
    elif confianca == "MEDIA":
        c_cell.fill = PatternFill("solid", fgColor="FEEBC8")
        c_cell.font = Font(color="7C2D12", bold=True)
    elif confianca == "BAIXA":
        c_cell.fill = PatternFill("solid", fgColor="FED7D7")
        c_cell.font = Font(color="742A2A", bold=True)

    ws.cell(row=idx, column=7, value=prazo if prazo else "")

    # Coluna "Mudou?" - destacar diferencas V1 vs V2
    mudou = "SIM" if regra_v1 != regra_v2 else ""
    m_cell = ws.cell(row=idx, column=8, value=mudou)
    if mudou == "SIM":
        m_cell.fill = PatternFill("solid", fgColor="FEF3C7")
        m_cell.font = Font(color="92400E", bold=True)
        m_cell.alignment = Alignment(horizontal="center")

    # Integra
    int_cell = ws.cell(row=idx, column=9, value=texto)
    int_cell.alignment = Alignment(wrap_text=True, vertical="top")

# Larguras das colunas
ws.column_dimensions["A"].width = 28  # Processo
ws.column_dimensions["B"].width = 12  # Data
ws.column_dimensions["C"].width = 15  # Natureza
ws.column_dimensions["D"].width = 38  # Regra V1
ws.column_dimensions["E"].width = 38  # Regra V2
ws.column_dimensions["F"].width = 11  # Confianca
ws.column_dimensions["G"].width = 8   # Prazo
ws.column_dimensions["H"].width = 10  # Mudou
ws.column_dimensions["I"].width = 100 # Integra

# Freeze primeira linha + 3 primeiras colunas
ws.freeze_panes = "D2"

# Filtro automatico
ws.auto_filter.ref = f"A1:I{len(pubs)+1}"

wb.save("intimacoes_djen.xlsx")
print(f"Planilha gerada: intimacoes_djen.xlsx ({len(pubs)} linhas)")
